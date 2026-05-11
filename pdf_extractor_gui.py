#!/usr/bin/env python3
"""
PDF 公文資訊擷取工具 - 圖形介面版
"""

import os
import re
import glob
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pdfplumber
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


FIELD_PATTERNS = {
    "主旨": [
        r"主\s*旨[：:]\s*(.*?)(?=\n\s*(?:說明|辦理|正本|副本|發文|附件|$))",
        r"主\s*旨[：:]\s*(.+)",
    ],
    "發文機關": [
        r"發\s*文\s*機\s*關[：:]\s*(.+)",
        r"機\s*關\s*名\s*稱[：:]\s*(.+)",
    ],
    "發文字號": [
        r"發\s*文\s*字\s*號[：:]\s*(.+)",
        r"字\s*號[：:]\s*(.+)",
    ],
}


def extract_text_from_pdf(pdf_path):
    text = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except Exception:
        try:
            reader = PdfReader(pdf_path)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        except Exception:
            pass
    return text


def extract_field(text, field_name):
    patterns = FIELD_PATTERNS.get(field_name, [])
    for pattern in patterns:
        match = re.search(pattern, text, re.DOTALL | re.MULTILINE)
        if match:
            value = match.group(1).strip()
            value = re.sub(r"\s+", " ", value).strip()
            if len(value) > 500:
                value = value[:500] + "..."
            return value
    return ""


def process_pdfs(pdf_files, output_path, log_fn, progress_fn):
    wb = Workbook()
    ws = wb.active
    ws.title = "公文資訊"

    header_font = Font(name="微軟正黑體", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", start_color="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["序號", "檔案名稱", "發文機關", "發文字號", "主旨"]
    col_widths = [6, 30, 25, 22, 60]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 24

    alt_fill = PatternFill("solid", start_color="EBF3FB")
    normal_fill = PatternFill("solid", start_color="FFFFFF")
    total = len(pdf_files)

    for idx, pdf_path in enumerate(pdf_files, 1):
        filename = os.path.basename(pdf_path)
        log_fn(f"[{idx}/{total}] 處理中：{filename}")
        progress_fn(idx, total)

        text = extract_text_from_pdf(pdf_path)
        agency = extract_field(text, "發文機關")
        doc_num = extract_field(text, "發文字號")
        subject = extract_field(text, "主旨")

        row = idx + 1
        row_fill = alt_fill if idx % 2 == 0 else normal_fill
        values = [idx, filename, agency, doc_num, subject]
        aligns = [center_align, wrap_align, wrap_align, wrap_align, wrap_align]

        for col_idx, (val, align) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = Font(name="微軟正黑體", size=10)
            cell.alignment = align
            cell.fill = row_fill
            cell.border = thin_border
        ws.row_dimensions[row].height = 40

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    wb.save(output_path)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PDF 公文資訊擷取工具")
        self.resizable(False, False)
        self.configure(bg="#F0F4F8")
        self._build_ui()
        self._center()

    def _center(self):
        self.update_idletasks()
        w, h = 620, 520
        x = (self.winfo_screenwidth() - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _build_ui(self):
        BLUE = "#1F4E79"
        LIGHT = "#F0F4F8"

        # 標題
        tk.Label(self, text="📄 PDF 公文資訊擷取工具",
                 font=("微軟正黑體", 16, "bold"),
                 bg=BLUE, fg="white", pady=14).pack(fill="x")

        tk.Label(self, text="自動擷取「主旨」「發文機關」「發文字號」並輸出至 Excel",
                 font=("微軟正黑體", 10), bg=BLUE, fg="#AED6F1", pady=4).pack(fill="x")

        main = tk.Frame(self, bg=LIGHT, padx=24, pady=16)
        main.pack(fill="both", expand=True)

        # PDF 來源
        tk.Label(main, text="選擇 PDF 檔案或資料夾", font=("微軟正黑體", 10, "bold"),
                 bg=LIGHT, fg="#2C3E50").grid(row=0, column=0, sticky="w", pady=(0, 4))

        src_frame = tk.Frame(main, bg=LIGHT)
        src_frame.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        src_frame.columnconfigure(0, weight=1)

        self.src_var = tk.StringVar()
        tk.Entry(src_frame, textvariable=self.src_var, font=("微軟正黑體", 10),
                 width=46, relief="solid", bd=1).grid(row=0, column=0, ipady=5, sticky="ew")

        btn_frame = tk.Frame(src_frame, bg=LIGHT)
        btn_frame.grid(row=0, column=1, padx=(8, 0))
        tk.Button(btn_frame, text="選擇檔案", command=self._pick_files,
                  bg="#2980B9", fg="white", font=("微軟正黑體", 9),
                  relief="flat", padx=8, pady=4, cursor="hand2").pack(side="left", padx=(0, 4))
        tk.Button(btn_frame, text="選擇資料夾", command=self._pick_folder,
                  bg="#2980B9", fg="white", font=("微軟正黑體", 9),
                  relief="flat", padx=8, pady=4, cursor="hand2").pack(side="left")

        # 輸出路徑
        tk.Label(main, text="輸出 Excel 儲存位置", font=("微軟正黑體", 10, "bold"),
                 bg=LIGHT, fg="#2C3E50").grid(row=2, column=0, sticky="w", pady=(0, 4))

        out_frame = tk.Frame(main, bg=LIGHT)
        out_frame.grid(row=3, column=0, sticky="ew", pady=(0, 16))
        out_frame.columnconfigure(0, weight=1)

        self.out_var = tk.StringVar(value=os.path.join(os.path.expanduser("~"), "Desktop", "公文資訊整理.xlsx"))
        tk.Entry(out_frame, textvariable=self.out_var, font=("微軟正黑體", 10),
                 width=46, relief="solid", bd=1).grid(row=0, column=0, ipady=5, sticky="ew")
        tk.Button(out_frame, text="瀏覽", command=self._pick_output,
                  bg="#2980B9", fg="white", font=("微軟正黑體", 9),
                  relief="flat", padx=8, pady=4, cursor="hand2").grid(row=0, column=1, padx=(8, 0))

        # 執行按鈕
        self.run_btn = tk.Button(main, text="▶  開始擷取", command=self._run,
                                  bg=BLUE, fg="white", font=("微軟正黑體", 12, "bold"),
                                  relief="flat", pady=10, cursor="hand2")
        self.run_btn.grid(row=4, column=0, sticky="ew", pady=(0, 14))

        # 進度條
        self.progress = ttk.Progressbar(main, mode="determinate", length=560)
        self.progress.grid(row=5, column=0, sticky="ew", pady=(0, 8))

        self.status_var = tk.StringVar(value="請選擇 PDF 檔案或資料夾後按下「開始擷取」")
        tk.Label(main, textvariable=self.status_var, font=("微軟正黑體", 9),
                 bg=LIGHT, fg="#555", anchor="w").grid(row=6, column=0, sticky="w")

        # 日誌
        tk.Label(main, text="處理記錄", font=("微軟正黑體", 10, "bold"),
                 bg=LIGHT, fg="#2C3E50").grid(row=7, column=0, sticky="w", pady=(12, 4))

        log_frame = tk.Frame(main, bg=LIGHT)
        log_frame.grid(row=8, column=0, sticky="nsew")
        main.rowconfigure(8, weight=1)
        main.columnconfigure(0, weight=1)

        self.log = tk.Text(log_frame, height=8, font=("Consolas", 9),
                           relief="solid", bd=1, state="disabled",
                           bg="#1A1A2E", fg="#A8D8EA", insertbackground="white")
        scroll = tk.Scrollbar(log_frame, command=self.log.yview)
        self.log.configure(yscrollcommand=scroll.set)
        self.log.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        self._pdf_files = []

    def _pick_files(self):
        files = filedialog.askopenfilenames(
            title="選擇 PDF 檔案", filetypes=[("PDF 檔案", "*.pdf"), ("所有檔案", "*.*")]
        )
        if files:
            self._pdf_files = list(files)
            self.src_var.set(f"已選擇 {len(files)} 個檔案")
            self._log(f"已選擇 {len(files)} 個 PDF 檔案")

    def _pick_folder(self):
        folder = filedialog.askdirectory(title="選擇包含 PDF 的資料夾")
        if folder:
            files = glob.glob(os.path.join(folder, "**/*.pdf"), recursive=True)
            files += glob.glob(os.path.join(folder, "*.pdf"))
            self._pdf_files = list(set(files))
            self.src_var.set(folder)
            self._log(f"資料夾：{folder}，找到 {len(self._pdf_files)} 個 PDF")

    def _pick_output(self):
        path = filedialog.asksaveasfilename(
            title="儲存 Excel 檔案",
            defaultextension=".xlsx",
            filetypes=[("Excel 檔案", "*.xlsx")],
            initialfile="公文資訊整理.xlsx",
        )
        if path:
            self.out_var.set(path)

    def _log(self, msg):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")

    def _run(self):
        if not self._pdf_files:
            messagebox.showwarning("提示", "請先選擇 PDF 檔案或資料夾！")
            return
        out = self.out_var.get().strip()
        if not out:
            messagebox.showwarning("提示", "請指定輸出 Excel 路徑！")
            return

        self.run_btn.config(state="disabled", text="處理中…")
        self.progress["value"] = 0
        self._log(f"\n{'='*40}")
        self._log(f"開始處理 {len(self._pdf_files)} 個檔案…")

        def worker():
            try:
                process_pdfs(
                    self._pdf_files, out,
                    log_fn=lambda m: self.after(0, self._log, m),
                    progress_fn=lambda i, t: self.after(0, self._set_progress, i, t),
                )
                self.after(0, self._done, out)
            except Exception as e:
                self.after(0, self._error, str(e))

        threading.Thread(target=worker, daemon=True).start()

    def _set_progress(self, current, total):
        pct = int(current / total * 100)
        self.progress["value"] = pct
        self.status_var.set(f"處理中… {current}/{total} ({pct}%)")

    def _done(self, out):
        self.run_btn.config(state="normal", text="▶  開始擷取")
        self.progress["value"] = 100
        self.status_var.set(f"✅ 完成！已儲存至：{out}")
        self._log(f"\n✅ 完成！Excel 已儲存至：\n   {out}")
        messagebox.showinfo("完成", f"擷取完成！\n\nExcel 已儲存至：\n{out}")

    def _error(self, msg):
        self.run_btn.config(state="normal", text="▶  開始擷取")
        self.status_var.set("❌ 發生錯誤")
        self._log(f"❌ 錯誤：{msg}")
        messagebox.showerror("錯誤", msg)


if __name__ == "__main__":
    app = App()
    app.mainloop()
